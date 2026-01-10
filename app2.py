import streamlit as st  # 匯入 Streamlit 工具箱，用來製作網頁介面
import yfinance as yf  # 匯入 Yahoo Finance 工具，用來下載全球股市和加密貨幣的歷史股價
import pandas as pd  # 匯入 Pandas 工具
import numpy as np  # 匯入 Numpy 工具
import requests  # 匯入 Requests 工具
import re  # 匯入正規表示法工具
from io import BytesIO  # 匯入記憶體緩存工具
from openpyxl import Workbook  # 匯入製作 Excel 的工具
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side  # 匯入 Excel 樣式工具
import datetime  # 匯入日期時間工具

# --- FinMind API 配置 ---
FINMIND_TOKEN = "eyJ0eXAiOiJKV1QiLCJhbGciOiJIUzI1NiJ9.eyJkYXRlIjoiMjAyNi0wMS0xMCAxMzo1NzoyMCIsInVzZXJfaWQiOiJWaXNpb24iLCJlbWFpbCI6ImRlbGlnaHRpbnRoZWtva0BnbWFpbC5jb20iLCJpcCI6IjEuMTcxLjIwNy4xMzgifQ.Eo7VFBIK50f_PWWIvBMnskSSYNGihrcC0nS1SHmjAdQ"

def get_finmind_data(dataset, stock_id, start_date):
    url = "https://api.finmindtrade.com/api/v4/data"
    headers = {"Authorization": f"Bearer {FINMIND_TOKEN}"}
    parameter = {"dataset": dataset, "data_id": stock_id, "start_date": start_date}
    try:
        resp = requests.get(url, params=parameter, headers=headers, timeout=10)
        res_json = resp.json()
        if res_json.get("msg") == "success":
            return pd.DataFrame(res_json["data"])
    except:
        pass
    return pd.DataFrame()

# --- SuperTrend 計算 ---
def calculate_st_full(df, period, multiplier):
    df_st = df.copy().reset_index(drop=True)
    high, low, close = df_st['High'], df_st['Low'], df_st['Close']
    tr = pd.concat([high - low, abs(high - close.shift(1)), abs(low - close.shift(1))], axis=1).max(axis=1)
    atr = tr.rolling(period).mean()
    hl2 = (high + low) / 2
    f_upper, f_lower = hl2 + (multiplier * atr), hl2 - (multiplier * atr)
    direction = np.ones(len(df_st))
    ub, lb, c = f_upper.values, f_lower.values, close.values
    for i in range(period, len(df_st)):
        if c[i-1] > lb[i-1]: lb[i] = max(lb[i], lb[i-1])
        if c[i-1] < ub[i-1]: ub[i] = min(ub[i], ub[i-1])
        if i < len(df_st) and c[i] > ub[i-1]:
            direction[i] = 1
        elif i < len(df_st) and c[i] < lb[i-1]:
            direction[i] = -1
        else:
            direction[i] = direction[i-1]
        if direction[i] == 1 and lb[i] < lb[i-1]: lb[i] = lb[i-1]
        if direction[i] == -1 and ub[i] > ub[i-1]: ub[i] = ub[i-1]
    return direction, ub, lb

# --- UI 配置 ---
st.set_page_config(page_title="Fish Diagnoser E1.4.5", layout="wide")
st.title("盛夏風情・魚兒診斷器 (E1.4.5)")

# --- Sidebar ---
st.sidebar.header("🔍 診斷參數設定")
lookback = st.sidebar.selectbox("追溯參考天數", [3, 5, 10, 20, 60], index=2)
st.sidebar.header("🥢 SuperTrend 參數")
long_p, long_m = st.sidebar.number_input("長期 ATR 週期", value=120), st.sidebar.number_input("長期系數", value=4.0)
short_p, short_m = st.sidebar.number_input("短期 ATR 週期", value=3), st.sidebar.number_input("短期系數", value=2.0)

query = st.text_area("🐟 輸入代碼 (例如: 1609, 2330, btc)", height=100)

if query:
    input_list = [t.strip().upper() for t in query.replace(',', ' ').split() if t.strip()]
    input_list = ["BTC-USD" if x == "BTC" else x for x in input_list]

    if input_list:
        selected_tickers = []
        st.subheader("📌 請確認診斷對象")
        cols = st.columns(min(len(input_list), 3))
        for idx, q in enumerate(input_list):
            with cols[idx % 3]:
                search_res = yf.Search(q, max_results=5).quotes
                if search_res:
                    search_res = sorted(search_res, key=lambda x: 0 if x['symbol'].endswith(('.TW', '.TWO')) else 1)
                    options = {f"{r['symbol']} ({r.get('longname', '未知')})": r['symbol'] for r in search_res}
                    chosen = st.selectbox(f"搜尋 '{q}'：", list(options.keys()), key=f"sel_{q}_{idx}")
                    selected_tickers.append((options[chosen], chosen))

        if st.button("🚀 開始完整診斷", use_container_width=True):
            results_for_excel = []
            for idx, (target_ticker, display_name) in enumerate(selected_tickers):
                try:
                    raw = yf.download(target_ticker, period="2y", progress=False)
                    if not raw.empty:
                        df = raw.copy()
                        if isinstance(df.columns, pd.MultiIndex):
                            df.columns = df.columns.get_level_values(0)
                        curr_p = float(df['Close'].iloc[-1])
                        base_p = float(df['Close'].iloc[-(lookback + 1)])
                        ma5 = df['Close'].rolling(5).mean().iloc[-1]
                        ma10 = df['Close'].rolling(10).mean().iloc[-1]
                        ma20 = df['Close'].rolling(20).mean().iloc[-1]
                        ma60 = df['Close'].rolling(60).mean().iloc[-1]
                        ma120 = df['Close'].rolling(120).mean().iloc[-1]
                        ema20 = df['Close'].ewm(span=20, adjust=False).mean().iloc[-1]
                        curr_vol = df['Volume'].iloc[-1]
                        vol_ma5 = df['Volume'].rolling(5).mean().iloc[-1]
                        l_dir, _, _ = calculate_st_full(df, long_p, long_m)
                        s_dir, _, _ = calculate_st_full(df, short_p, short_m)
                        status_map = {(1, 1): ("✨ 浮光躍金", "#FFD700"), (-1, 1): ("🚀 靈魚突圍", "#00FFFF"), (1, -1): ("🍂 迴游潛歇", "#FFA500")}
                        final_label, status_color = status_map.get((l_dir[-1], s_dir[-1]), ("🌑 影跡稀微", "#A9A9A9"))

                        results_for_excel.append([target_ticker, curr_p, ma5, ma10, ma20, ma60, ma120, ema20, curr_vol, vol_ma5, final_label, s_dir[-1], l_dir[-1], base_p])

                        with st.expander(f"🔍 {display_name} - {final_label}", expanded=False):
                            p_pct = ((curr_p - base_p) / base_p) * 100
                            st.markdown(f"""
                            <div style="display: flex; justify-content: space-between; padding: 15px; background-color: #1e1e1e; border-radius: 10px; border: 1px solid #333; margin-bottom: 20px;">
                            <div style="flex: 1;"><div style="color: #aaa; font-size: 0.9rem;">目前現價</div><div style="font-size: 1.8rem; font-weight: bold; color: white;">{curr_p:,.2f}</div></div>
                            <div style="flex: 1;"><div style="color: #aaa; font-size: 0.9rem;">{lookback}日漲跌</div><div style="font-size: 1.8rem; font-weight: bold; color: white;">{p_pct:+.2f}%</div></div>
                            <div style="flex: 1;"><div style="color: #aaa; font-size: 0.9rem;">綜合判定</div><div style="font-size: 1.6rem; font-weight: bold; color: {status_color};">{final_label}</div></div>
                            </div>
                            """, unsafe_allow_html=True)
                            
                            col_bull, col_bear = st.columns(2)
                            red_check = '<span style="color:#FF4B4B; font-weight:bold;">✔</span>'
                            with col_bull:
                                st.markdown("### 🟠 多方動能")
                                if curr_p > ma60: st.markdown(f'<div style="padding:10px; border-radius:5px; background-color:rgba(255,140,0,0.1); border-left:5px solid #FF8C00; color:white; margin-bottom:10px;">{red_check} 生命線：守穩 60MA 之上</div>', unsafe_allow_html=True)
                                if ma20 > ma60: st.markdown(f'<div style="padding:10px; border-radius:5px; background-color:rgba(255,140,0,0.1); border-left:5px solid #FF8C00; color:white; margin-bottom:10px;">{red_check} 中長期趨勢：20MA/60MA 黃金交叉</div>', unsafe_allow_html=True)
                                if curr_p >= ma20: st.markdown(f'<div style="padding:10px; border-radius:5px; background-color:rgba(255,140,0,0.1); border-left:5px solid #FF8C00; color:white; margin-bottom:10px;">{red_check} 位階判定：目前站穩月線</div>', unsafe_allow_html=True)
                                if s_dir[-1] == 1: st.markdown(f'<div style="padding:10px; border-radius:5px; background-color:rgba(255,140,0,0.1); border-left:5px solid #FF8C00; color:white; margin-bottom:10px;">{red_check} SuperTrend：短線維持多頭</div>', unsafe_allow_html=True)
                                if l_dir[-1] == 1: st.markdown(f'<div style="padding:10px; border-radius:5px; background-color:rgba(255,140,0,0.1); border-left:5px solid #FF8C00; color:white; margin-bottom:10px;">{red_check} SuperTrend：長線背景偏多</div>', unsafe_allow_html=True)
                            with col_bear:
                                st.markdown("### 🔵 空方警示")
                                if curr_p < ma60: st.markdown('<div style="padding:10px; border-radius:5px; background-color:rgba(30,144,255,0.1); border-left:5px solid #1E90FF; color:white; margin-bottom:10px;">❌ 跌破 60MA 生命線</div>', unsafe_allow_html=True)
                                if ma20 < ma60: st.markdown('<div style="padding:10px; border-radius:5px; background-color:rgba(30,144,255,0.1); border-left:5px solid #1E90FF; color:white; margin-bottom:10px;">❌ 20MA/60MA 中長期死叉</div>', unsafe_allow_html=True)
                                if curr_p < ma20: st.markdown('<div style="padding:10px; border-radius:5px; background-color:rgba(30,144,255,0.1); border-left:5px solid #1E90FF; color:white; margin-bottom:10px;">❌ 位階偏低：目前在月線下</div>', unsafe_allow_html=True)
                                if s_dir[-1] == -1: st.markdown('<div style="padding:10px; border-radius:5px; background-color:rgba(30,144,255,0.1); border-left:5px solid #1E90FF; color:white; margin-bottom:10px;">❌ SuperTrend：短線轉弱</div>', unsafe_allow_html=True)
                                if l_dir[-1] == -1: st.markdown('<div style="padding:10px; border-radius:5px; background-color:rgba(30,144,255,0.1); border-left:5px solid #1E90FF; color:white; margin-bottom:10px;">❌ SuperTrend：長線背景偏空</div>', unsafe_allow_html=True)

                            st.markdown("---")
                            sid_only = re.sub(r'\D', '', target_ticker)
                            if "USD" not in target_ticker:
                                col_f1, col_f2 = st.columns(2)
                                with col_f1:
                                    st.write("📊 **營收精算 (月份偏移校正)**")
                                    rev_df = get_finmind_data("TaiwanStockMonthRevenue", sid_only, "2023-11-01")
                                    if not rev_df.empty:
                                        rev_df = rev_df[rev_df['revenue'] > 0].sort_values(by='date', ascending=True).reset_index(drop=True)
                                        rev_df['MoM'], rev_df['YoY'] = rev_df['revenue'].pct_change() * 100, rev_df['revenue'].pct_change(12) * 100
                                        for _, r in rev_df.tail(3).sort_values(by='date', ascending=False).iterrows():
                                            m = (pd.to_datetime(r['date']) - pd.DateOffset(months=1)).strftime('%m')
                                            # --- 修改 1 (Line 114): 修改營收 MoM/YoY 紅綠顯色邏輯 ---
                                            mom_c = "#FF4B4B" if r['MoM'] > 0 else "#00A000"
                                            yoy_c = "#FF4B4B" if r['YoY'] > 0 else "#00A000"
                                            st.write(f"**{m}月營收**：{r['revenue']/1e8:,.2f} 億 | MoM: <span style='color:{mom_c}; font-weight:bold;'>{r['MoM']:+.1f}%</span> | YoY: <span style='color:{yoy_c}; font-weight:bold;'>{r['YoY']:+.1f}%</span>", unsafe_allow_html=True)
                                with col_f2:
                                    st.write("💰 **最新季報 EPS**")
                                    eps_df = get_finmind_data("TaiwanStockFinancialStatements", sid_only, "2025-01-01")
                                    if not eps_df.empty:
                                        for _, r in eps_df[eps_df['type'] == 'EPS'].tail(3).sort_values(by='date', ascending=False).iterrows():
                                            dt = pd.to_datetime(r['date'])
                                            # --- 修改 2 (Line 120-121): 修改 EPS 正負紅綠顯色邏輯 ---
                                            eps_c = "#FF4B4B" if r['value'] > 0 else "#00A000"
                                            st.write(f"**{dt.year} Q{((dt.month-1)//3)+1} EPS**：<span style='color:{eps_c}; font-weight:bold;'>{r['value']:.2f}</span> 元", unsafe_allow_html=True)
                            tv_p = "BINANCE" if "USD" in target_ticker else ("TPEX" if ".TWO" in target_ticker else "TWSE")
                            tv_c = "BTCUSD" if "BTC" in target_ticker else sid_only
                            st.markdown(f"[🔗 開啟 TradingView 詳細圖表](https://www.tradingview.com/chart/?symbol={tv_p}:{tv_c})")

                except Exception as e:
                    st.error(f"分析失敗: {e}")

            # --- Excel 專業匯出 ---
            if results_for_excel:
                output = BytesIO()
                wb = Workbook()
                ws = wb.active
                ws.title = "Fish_Diagnosis"
                f14, f14b = Font(name='Calibri', size=14), Font(name='Calibri', size=14, bold=True)
                f_org, f_blu = PatternFill("solid", fgColor="FCD5B4"), PatternFill("solid", fgColor="DDEBF7")
                brd = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                aln_center = Alignment(horizontal='center', vertical='center', wrapText=True)

                params_list = [
                    ["入攤時間", datetime.datetime.now().strftime("%Y-%m-%d %H:%M")],
                    ["水波餘紋", f"{lookback}日參考"],
                    ["長期 ATR 設定", f"週期: {long_p} / 系數: {long_m}"],
                    ["短期 ATR 設定", f"週期: {short_p} / 系數: {short_m}"]
                ]
                for r, (k, v) in enumerate(params_list, 1):
                    for c, val in enumerate([k, v], 1):
                        cell = ws.cell(r, c, val)
                        cell.font, cell.fill, cell.border, cell.alignment = f14b if c==1 else f14, f_org, brd, aln_center

                header_row = 10
                h_list = ["序號", "個股代碼", "目前現價", "對比漲跌幅", "多方動能項目", "空方警示項目", "最新月營收 (M/Y)", "最新季報 EPS", "綜合判定"]
                for c, h in enumerate(h_list, 1):
                    cell = ws.cell(header_row, c, h)
                    cell.font, cell.fill, cell.border, cell.alignment = f14b, f_blu, brd, aln_center

                for r_idx, row_raw in enumerate(results_for_excel, header_row + 1):
                    sid, cp, m20, m60, s_st, l_st, bp = row_raw[0], row_raw[1], row_raw[4], row_raw[5], row_raw[11], row_raw[12], row_raw[13]
                    bull_items = "、".join([t for t, c in zip(["守穩60MA","20/60MA金叉","站穩月線","ST短多","ST長多"], [cp>m60, m20>m60, cp>=m20, s_st==1, l_st==1]) if c])
                    bear_items = "、".join([t for t, c in zip(["破60MA","20/60MA死叉","月線下","ST短轉弱","ST長轉弱"], [cp<m60, m20<m60, cp<m20, s_st==-1, l_st==-1]) if c])
                    
                    rev_summary, eps_summary = "N/A", "N/A"
                    sid_clean = re.sub(r'\D', '', sid)
                    if "USD" not in sid:
                        r_df = get_finmind_data("TaiwanStockMonthRevenue", sid_clean, "2023-11-01")
                        if not r_df.empty:
                            r_df = r_df[r_df['revenue']>0].sort_values(by='date', ascending=True).reset_index(drop=True)
                            r_df['MoM'], r_df['YoY'] = r_df['revenue'].pct_change()*100, r_df['revenue'].pct_change(12)*100
                            last = r_df.iloc[-1]
                            m = (pd.to_datetime(last['date']) - pd.DateOffset(months=1)).strftime('%m')
                            rev_summary = f"{m}月:{last['revenue']/1e8:.1f}億 (M:{last['MoM']:+.1f}% / Y:{last['YoY']:+.1f}%)"
                        e_df = get_finmind_data("TaiwanStockFinancialStatements", sid_clean, "2025-01-01")
                        if not e_df.empty:
                            last_e = e_df[e_df['type'] == 'EPS'].sort_values(by='date', ascending=True).iloc[-1]
                            eps_summary = f"{pd.to_datetime(last_e['date']).year}Q{((pd.to_datetime(last_e['date']).month-1)//3)+1}: {last_e['value']:.2f}元"

                    row_vals = [r_idx-header_row, sid, round(cp, 2), f"{((cp-bp)/bp)*100:+.2f}%", bull_items if bull_items else "無", bear_items if bear_items else "無", rev_summary, eps_summary, row_raw[10]]
                    for c_idx, val in enumerate(row_vals, 1):
                        cell = ws.cell(r_idx, c_idx, val)
                        cell.font, cell.fill, cell.border, cell.alignment = f14, f_blu, brd, aln_center

                for i, w in enumerate([15, 30, 15, 15, 50, 50, 45, 25, 20], 1):
                    ws.column_dimensions[ws.cell(header_row, i).column_letter].width = w

                wb.save(output)
                st.download_button(label="📥 下載 Excel 專業診斷報表", data=output.getvalue(), file_name=f"魚兒健檢_{datetime.date.today()}.xlsx", use_container_width=True)
